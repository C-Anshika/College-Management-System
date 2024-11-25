import re
from openpyxl import Workbook, load_workbook

def validate_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))

def load_or_create_workbook(file_name, sheet_name, headers):
    try:
        workbook = load_workbook(file_name)
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)
            workbook.save(file_name)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        workbook.save(file_name)
    return workbook
